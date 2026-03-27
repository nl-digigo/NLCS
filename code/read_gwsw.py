import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\100289\Downloads\GWSW uitgeschreven.xlsx')
ws = wb.active

# Get all unique categories
categories = set()
for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 1).value
    if cat:
        categories.add(str(cat).strip())

print("Unique categories:", sorted(categories))
print()

# Print all rows with category "Vraag"
print("=== ROWS WITH CATEGORY 'Vraag' ===")
for row in range(2, ws.max_row + 1):
    cat = str(ws.cell(row, 1).value or '').strip()
    if cat == 'Vraag':
        opmerking = ws.cell(row, 2).value or ''
        mapping = ws.cell(row, 3).value or ''
        mapping_sym = ws.cell(row, 4).value or ''
        soort = ws.cell(row, 5).value or ''
        fysiek_cols = []
        for c in range(6, 12):
            v = ws.cell(row, c).value
            if v:
                fysiek_cols.append(str(v))
        fysiek = ' > '.join(fysiek_cols)
        print(f"ROW {row}: Opmerking: {opmerking}")
        print(f"  Mapping NLCS: {mapping} | Symbool: {mapping_sym}")
        print(f"  Soort: {soort} | Fysiek: {fysiek}")
        print()

# Also print all unique opmerkingen per category
print("\n=== ALL UNIQUE OPMERKINGEN ===")
opmerkingen_by_cat = {}
for row in range(2, ws.max_row + 1):
    cat = str(ws.cell(row, 1).value or '').strip()
    opmerking = str(ws.cell(row, 2).value or '').strip()
    if opmerking:
        if cat not in opmerkingen_by_cat:
            opmerkingen_by_cat[cat] = set()
        opmerkingen_by_cat[cat].add(opmerking)

for cat in sorted(opmerkingen_by_cat.keys()):
    print(f"\n--- {cat} ---")
    for opm in sorted(opmerkingen_by_cat[cat]):
        print(f"  {opm}")
